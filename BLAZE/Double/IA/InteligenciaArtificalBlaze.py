import requests
import time
from datetime import datetime
from openpyxl import *
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.ensemble import ExtraTreesClassifier

def QualCor(aux):
    if  aux == 0: return 'BRANCO'
    elif aux == 1: return 'VERMELHO'
    elif aux == 2: return 'PRETO'



url = 'https://blaze.com/api/roulette_games/recent'

ultimaCor = ''
parcialWin = 0
parcialLoss = 0
auxMaxLoss = 0
auxMaxWin = 0
maxLoss = 0
maxWin = 0
while True:
    df = pd.read_excel('bd.xlsx')
    dadosAtual = pd.DataFrame(df.loc[(df['COR_PREVISTA'] != 0) & (df['COR_PREVISTA'] != 1) & (df['COR_PREVISTA'] != 2)])
    dadosAtual = dadosAtual.drop('COR_PREVISTA', axis=1)
    df = df.drop(int(len(df) - 1))
    y = df['COR_PREVISTA']
    x = df.drop('COR_PREVISTA', axis=1)
    # dadosAtual = x[-1::]
    x_treino, x_teste, y_treino, y_teste = train_test_split(x,y)
    modelo = ExtraTreesClassifier()
    modelo.fit(x_treino,y_treino)
    resultado = modelo.score(x_teste,y_teste)
    print(f'\n\nAcurácia Atual: {round(100*resultado,2)}')    
    try:
        response = requests.get(url)
        r = response.json()
    except:pass
    hora = float(((datetime.now()).strftime('%H')))
    minuto = float(((datetime.now()).strftime('%M')))
    try: 
        wb = load_workbook(filename='bd.xlsx')
        sheet = wb.active
        count = 3
        for rows in sheet.iter_rows(min_row=2):
            count += 1
        count -= 1
        sheet.cell(row=count, column=21).value = minuto
        sheet.cell(row=count, column=20).value = hora
        sheet.cell(row=count, column=19).value = r[0]['color']
        sheet.cell(row=count, column=18).value = r[1]['color']
        sheet.cell(row=count, column=17).value = r[2]['color']
        sheet.cell(row=count, column=16).value = r[3]['color']
        sheet.cell(row=count, column=15).value = r[4]['color']
        sheet.cell(row=count, column=14).value = r[5]['color']
        sheet.cell(row=count, column=13).value = r[6]['color']
        sheet.cell(row=count, column=12).value = r[7]['color']
        sheet.cell(row=count, column=11).value = r[8]['color']
        sheet.cell(row=count, column=10).value = r[9]['color']
        sheet.cell(row=count, column=9).value = r[10]['color']
        sheet.cell(row=count, column=8).value = r[11]['color']
        sheet.cell(row=count, column=7).value = r[12]['color']
        sheet.cell(row=count, column=6).value = r[13]['color']
        sheet.cell(row=count, column=5).value = r[14]['color']
        sheet.cell(row=count, column=4).value = r[15]['color']
        sheet.cell(row=count, column=3).value = r[16]['color']
        sheet.cell(row=count, column=2).value = r[17]['color']
        sheet.cell(row=count, column=1).value = r[18]['color']

        wb.save(filename='bd.xlsx')
        wb.close
        print('Dados Salvos')
    except: pass
    corPrevista = QualCor(modelo.predict(dadosAtual))
    print(f'Previsão cor: {corPrevista}') 
        
    idAnterior = r[0]['id']
    while idAnterior == r[0]['id']:
        try:
            response = requests.get(url)
            r = response.json()
        except: pass
    
    try:     
        wb = load_workbook(filename='dadosBlaze.xlsx')
        sheet = wb.active
        count = 2
        for rows in sheet.iter_rows(min_row=2):
            count += 1
        count -=1
        sheet.cell(row=count, column=22).value = r[0]['color']
        wb.save(filename='dadosBlaze.xlsx')
        wb.close
    except: pass        
            
    if corPrevista != '':
        if QualCor(r[0]['color']) == corPrevista:
            auxMaxLoss = 0
            parcialWin += 1
            auxMaxWin += 1
            if maxWin < auxMaxWin: maxWin = auxMaxWin
            print('-> ✅ Green')
        else:
            auxMaxWin = 0
            parcialLoss += 1
            auxMaxLoss += 1
            if maxLoss < auxMaxLoss: maxLoss = auxMaxLoss
            print('-> ⛔️ Loss')
        winrate = round(100*(parcialWin / (parcialWin + parcialLoss)),2)
        msg = f'-> Parcial: ✅ = {parcialWin} | ⛔️ = {parcialLoss} | 🎯 = {winrate}% de acerto'
        print(msg)
        print(f'-> Máximo de VITORIA seguida: {maxWin}')
        print(f'-> Máximo de DERROTA seguida: {maxLoss}')
    
    # break
    
    
    