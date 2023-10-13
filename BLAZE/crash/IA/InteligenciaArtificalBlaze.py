import requests
import time
from datetime import datetime
from openpyxl import *
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.ensemble import ExtraTreesClassifier

def Send_Message_Telegram(message):
    token = '5963411116:AAHeovmtOeDh0-g-L6R6KX4IjoyzP7nHam8'
    chat_id = '-790930643'
    try:
        data = {"chat_id": chat_id, "text": message}
        urlTelegram = "https://api.telegram.org/bot{}/sendMessage".format(token)
        requests.post(urlTelegram, data)
    except Exception as e:
        print("Erro no sendMessage:", e)

url = 'https://blaze.com/api/crash_games/recent'
conatodorRodadas = 0
baseDados = 'dadosBlaze10.xlsx'
multiplicador = 9.99
while True:
    if conatodorRodadas == 0:
        conatodorRodadas = 10
        try:
            arquivo = pd.read_excel(baseDados)
            y = arquivo['COR_PREVISTA']
            x = arquivo.drop('COR_PREVISTA', axis=1)
            x_treino, x_teste, y_treino, y_teste = train_test_split(x,y,test_size = 0.3)
            modelo = ExtraTreesClassifier()
            modelo.fit(x_treino,y_treino)
            resultado = modelo.score(x_teste,y_teste)
            print(f'\n\nAcurácia Atual: {round(100*resultado,2)}')
            
        except: pass
        
    response = requests.get(url)
    r = response.json()
        

    try: 
        wb = load_workbook(filename=baseDados)
        sheet = wb.active
        count = 3
        for rows in sheet.iter_rows(min_row=2):count += 1
        count -= 1
        
        hora = float(((datetime.now()).strftime('%H')))
        minuto = float(((datetime.now()).strftime('%M')))
        if hora >= 0 and hora <=5:sheet.cell(row=count, column=15).value = 1
        elif hora >= 6 and hora <=11: sheet.cell(row=count, column=15).value = 2
        elif hora >= 12 and hora <=17: sheet.cell(row=count, column=15).value = 3
        elif hora >= 18 and hora <=23: sheet.cell(row=count, column=15).value = 4
        
        # if hora == 0 or hora == 6 or hora == 12 or hora == 18:
        #     sheet.cell(row=count, column=16).value = 1
        # elif hora == 1 or hora == 7 or hora == 13 or hora == 19:
        #     sheet.cell(row=count, column=16).value = 2
        # elif hora == 2 or hora == 8 or hora == 14 or hora == 20:
        #     sheet.cell(row=count, column=16).value = 3
        # elif hora == 3 or hora == 9 or hora == 15 or hora == 21:
        #     sheet.cell(row=count, column=16).value = 4
        # elif hora == 4 or hora == 10 or hora == 16 or hora == 22:
        #     sheet.cell(row=count, column=16).value = 5
        # elif hora == 5 or hora == 11 or hora == 17 or hora == 13:
        #     sheet.cell(row=count, column=16).value = 6
        sheet.cell(row=count, column=16).value = hora
        sheet.cell(row=count, column=17).value =  minuto  
        
        
        if r[0]['crash_point'] != '0': sheet.cell(row=count, column=14).value = r[0]['crash_point']
        else: sheet.cell(row=count, column=14).value = '1.00'
        
        if  r[1]['crash_point'] !='0': sheet.cell(row=count, column=13).value = r[1]['crash_point']
        else: sheet.cell(row=count, column=13).value = '1.00'
        
        if  r[2]['crash_point'] != '0': sheet.cell(row=count, column=12).value = r[2]['crash_point']
        else: sheet.cell(row=count, column=12).value = '1.00'
        
        if  r[3]['crash_point'] != '0': sheet.cell(row=count, column=11).value = r[3]['crash_point']
        else: sheet.cell(row=count, column=11).value = '1.00'
        
        if  r[4]['crash_point'] != '0': sheet.cell(row=count, column=10).value = r[4]['crash_point']
        else: sheet.cell(row=count, column=10).value = '1.00'
        
        if  r[5]['crash_point'] != '0': sheet.cell(row=count, column=9).value = r[5]['crash_point']
        else: sheet.cell(row=count, column=9).value = '1.00'
        
        if  r[6]['crash_point'] != '0':sheet.cell(row=count, column=8).value = r[6]['crash_point']
        else: sheet.cell(row=count, column=8).value = '1.00'
        
        if  r[7]['crash_point'] != '0': sheet.cell(row=count, column=7).value = r[7]['crash_point']
        else: sheet.cell(row=count, column=7).value = '1.00'
        
        if  r[8]['crash_point'] != '0': sheet.cell(row=count, column=6).value = r[8]['crash_point']
        else: sheet.cell(row=count, column=6).value = '1.00'
        
        if  r[9]['crash_point'] != '0': sheet.cell(row=count, column=5).value = r[9]['crash_point']
        else: sheet.cell(row=count, column=5).value = '1.00'
        
        if  r[10]['crash_point'] != '0': sheet.cell(row=count, column=4).value = r[10]['crash_point']
        else: sheet.cell(row=count, column=4).value = '1.00'
        
        if  r[11]['crash_point'] != '0': sheet.cell(row=count, column=3).value = r[11]['crash_point']
        else: sheet.cell(row=count, column=3).value = '1.00'
        
        if  r[12]['crash_point'] != '0': sheet.cell(row=count, column=2).value = r[12]['crash_point']
        else: sheet.cell(row=count, column=2).value = '1.00'
        
        if  r[13]['crash_point'] != '0': sheet.cell(row=count, column=1).value = r[13]['crash_point']
        else: sheet.cell(row=count, column=1).value = '1.00'
        wb.save(filename=baseDados)
        wb.close
        print('Dados Salvos')
    except: pass
    arquivo = pd.read_excel(baseDados)
    x = arquivo.drop('COR_PREVISTA', axis=1)
    dadosAtual = x[-1::]
    previsao = ''
    print(previsao)   
    previsao = modelo.predict(dadosAtual)
    if previsao == 1:
        print(f'Operar, sair no {multiplicador + 0.01}x')
    elif previsao == 0:
        print('Não operar')
    print(previsao)   
    idAnterior = r[0]['id']
    while idAnterior == r[0]['id']:
        try:
            response = requests.get(url)
            r = response.json()
        except: pass
            
            
    if float(r[0]['crash_point']) > multiplicador and previsao == 1 or float(r[0]['crash_point']) < multiplicador and previsao == 0:
        print('Acertou !')
    elif float(r[0]['crash_point']) < multiplicador and previsao == 1 or float(r[0]['crash_point']) > multiplicador and previsao == 0:
        print('Errou !')
    try:     
        wb = load_workbook(filename=baseDados)
        sheet = wb.active
        count = 2
        for rows in sheet.iter_rows(min_row=2):
            count += 1
        count -=1
        if float(r[0]['crash_point']) > multiplicador :
            aux = 1
        else:
            aux = 0
        sheet.cell(row=count, column=18).value = aux
        wb.save(filename=baseDados)
        wb.close
    except: pass
    
    conatodorRodadas -= 1
    
    